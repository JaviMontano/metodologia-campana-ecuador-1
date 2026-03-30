#!/usr/bin/env python3
"""
data_utils.py — Data manipulation utilities for MetodologIA
Handles CSV, JSON, YAML, and Excel data transformations.

Usage:
    from data_utils import load_data, to_markdown_table, merge_sources, validate_prices

Or standalone:
    python data_utils.py csv2json data.csv output.json
    python data_utils.py yaml2json catalog.yaml output.json
    python data_utils.py json2csv data.json output.csv
    python data_utils.py validate-prices services.yaml conditions.yaml
"""

import sys
import json
import csv
import re
from pathlib import Path
from typing import Union

# ── Loaders ──────────────────────────────────────────────────────────────────

def load_data(filepath: str) -> Union[dict, list]:
    """Load data from JSON, YAML, CSV, or XLSX."""
    p = Path(filepath)
    ext = p.suffix.lower()

    if ext == '.json':
        return json.loads(p.read_text(encoding='utf-8'))
    elif ext in ('.yaml', '.yml'):
        import yaml
        return yaml.safe_load(p.read_text(encoding='utf-8'))
    elif ext == '.csv':
        with open(filepath, newline='', encoding='utf-8') as f:
            return list(csv.DictReader(f))
    elif ext == '.xlsx':
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(h or '') for h in rows[0]]
        return [dict(zip(headers, [str(c or '') for c in row])) for row in rows[1:]]
    else:
        raise ValueError(f'Unsupported format: {ext}')


def save_data(data: Union[dict, list], filepath: str):
    """Save data to JSON, YAML, or CSV."""
    p = Path(filepath)
    ext = p.suffix.lower()

    if ext == '.json':
        p.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding='utf-8')
    elif ext in ('.yaml', '.yml'):
        import yaml
        p.write_text(yaml.dump(data, allow_unicode=True, default_flow_style=False), encoding='utf-8')
    elif ext == '.csv':
        if isinstance(data, list) and data:
            keys = data[0].keys()
            with open(filepath, 'w', newline='', encoding='utf-8') as f:
                writer = csv.DictWriter(f, fieldnames=keys)
                writer.writeheader()
                writer.writerows(data)
    print(f'Saved: {filepath}')


# ── Transformations ──────────────────────────────────────────────────────────

def to_markdown_table(data: list, columns: list = None) -> str:
    """Convert list of dicts to markdown table."""
    if not data:
        return ''
    cols = columns or list(data[0].keys())
    lines = ['| ' + ' | '.join(cols) + ' |']
    lines.append('| ' + ' | '.join(['---'] * len(cols)) + ' |')
    for row in data:
        cells = [str(row.get(c, '')) for c in cols]
        lines.append('| ' + ' | '.join(cells) + ' |')
    return '\n'.join(lines)


def flatten_nested(data: dict, prefix: str = '') -> dict:
    """Flatten nested dict: {'a': {'b': 1}} → {'a.b': 1}"""
    result = {}
    for k, v in data.items():
        key = f'{prefix}.{k}' if prefix else k
        if isinstance(v, dict):
            result.update(flatten_nested(v, key))
        else:
            result[key] = v
    return result


def merge_sources(*sources: list) -> list:
    """Merge multiple data sources by common keys."""
    merged = {}
    for source in sources:
        for item in source:
            key = item.get('slug') or item.get('id') or item.get('name', '')
            if key in merged:
                merged[key].update(item)
            else:
                merged[key] = dict(item)
    return list(merged.values())


# ── Price Validation ─────────────────────────────────────────────────────────

COP_PATTERN = re.compile(r'(?:COP\s*)?[\d.,]+')

def parse_cop(value: str) -> int:
    """Parse COP string to integer: 'COP 12.000.000' → 12000000"""
    if not value or value.strip() in ('N/A', '-', '[POR CONFIRMAR]', ''):
        return 0
    digits = re.sub(r'[^0-9]', '', str(value))
    return int(digits) if digits else 0


def validate_prices(services_path: str, conditions_path: str = None) -> list:
    """Validate price consistency across catalog files. Returns list of issues."""
    services = load_data(services_path)
    issues = []

    if isinstance(services, dict):
        items = services.get('services', services.get('items', []))
    else:
        items = services

    for item in items:
        slug = item.get('slug', 'unknown')
        pricing = item.get('pricing', item.get('prices', {}))

        if isinstance(pricing, dict):
            for segment, price in pricing.items():
                if isinstance(price, str) and '[POR CONFIRMAR]' in price:
                    issues.append({
                        'slug': slug,
                        'segment': segment,
                        'severity': 'WARNING',
                        'message': f'Price pending confirmation: {price}'
                    })
                elif isinstance(price, (int, float)) and price <= 0:
                    issues.append({
                        'slug': slug,
                        'segment': segment,
                        'severity': 'ERROR',
                        'message': f'Invalid price: {price}'
                    })

    if conditions_path:
        conditions = load_data(conditions_path)
        pc_items = conditions if isinstance(conditions, list) else conditions.get('por_confirmar', [])
        for pc in pc_items:
            issues.append({
                'slug': pc.get('affected_services', 'general'),
                'segment': 'N/A',
                'severity': 'INFO',
                'message': f'[POR CONFIRMAR] {pc.get("id", "")}: {pc.get("description", pc.get("item", ""))}'
            })

    return issues


# ── Statistics ───────────────────────────────────────────────────────────────

def summarize(data: list, group_by: str = None) -> dict:
    """Basic summary statistics for a dataset."""
    summary = {
        'total_records': len(data),
        'columns': list(data[0].keys()) if data else [],
    }
    if group_by and data:
        groups = {}
        for item in data:
            key = str(item.get(group_by, 'N/A'))
            groups[key] = groups.get(key, 0) + 1
        summary['groups'] = groups
    return summary


# ── CLI ──────────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 3:
        print('Usage:')
        print('  python data_utils.py csv2json input.csv output.json')
        print('  python data_utils.py yaml2json input.yaml output.json')
        print('  python data_utils.py json2csv input.json output.csv')
        print('  python data_utils.py validate-prices services.yaml [conditions.yaml]')
        sys.exit(1)

    cmd = sys.argv[1]

    if cmd == 'validate-prices':
        issues = validate_prices(sys.argv[2], sys.argv[3] if len(sys.argv) > 3 else None)
        for issue in issues:
            print(f'[{issue["severity"]}] {issue["slug"]}/{issue["segment"]}: {issue["message"]}')
        print(f'\nTotal issues: {len(issues)}')
    elif cmd in ('csv2json', 'yaml2json', 'json2csv'):
        data = load_data(sys.argv[2])
        save_data(data, sys.argv[3])
    else:
        print(f'Unknown command: {cmd}')
        sys.exit(1)


if __name__ == '__main__':
    main()
